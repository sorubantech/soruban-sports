"""
Generate QA Test Plan Excel from the markdown test plan.
Run: python docs/qa/generate-test-plan-excel.py
Output: docs/qa/soruban-sports-test-plan.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colors
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
SUB_HEADER_FONT = Font(name="Calibri", bold=True, size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
PASS_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
BLOCKED_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_VALIDATION = DataValidation(type="list", formula1='"PASS,FAIL,BLOCKED,SKIP,N/A"', allow_blank=True)
STATUS_VALIDATION.error = "Please select PASS, FAIL, BLOCKED, SKIP, or N/A"
STATUS_VALIDATION.errorTitle = "Invalid Status"
SEVERITY_VALIDATION = DataValidation(type="list", formula1='"Critical,High,Medium,Low,Cosmetic"', allow_blank=True)
PRIORITY_VALIDATION = DataValidation(type="list", formula1='"P1-Fix Now,P2-Next Sprint,P3-Backlog,P4-Nice to Have"', allow_blank=True)


def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def add_section_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def add_sub_header_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SUB_HEADER_FILL
    cell.font = SUB_HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ──────────────────────────────────────────────
# CREATE WORKBOOK
# ──────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ══════════════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = "1B4F72"

overview_headers = ["Metric", "Count", "Notes"]
set_col_widths(ws, [35, 15, 50])

row = 1
ws.merge_cells("A1:C1")
ws.cell(row=1, column=1, value="SORUBAN SPORTS — QA TEST PLAN").font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
row = 2
ws.cell(row=2, column=1, value="Document Version: 1.0  |  Date: 21 Mar 2026  |  Prepared By: Senior QA Analyst").font = Font(name="Calibri", size=10, color="666666")
row = 4
for i, h in enumerate(overview_headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header_row(ws, row, 3)

overview_data = [
    ("Total Screens/Pages", 71, "Across all 4 apps"),
    ("End-to-End Workflows", 13, "Cross-app test scenarios"),
    ("Individual Test Steps", "210+", "Detailed step-by-step"),
    ("UI/UX Checks", 21, "Theme, responsive, navigation"),
    ("Chart Validations", 16, "All chart rendering checks"),
    ("Data Consistency Checks", 9, "Cross-app data verification"),
    ("Known Limitations", 10, "Expected prototype limitations"),
]
for d in overview_data:
    row += 1
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    style_data_row(ws, row, 3)

row += 2
ws.cell(row=row, column=1, value="Apps Under Test").font = BOLD_FONT
row += 1
app_headers = ["#", "App", "Path"]
for i, h in enumerate(app_headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header_row(ws, row, 3, SUB_HEADER_FILL, SUB_HEADER_FONT)

apps = [
    (1, "Student & Parent App", "prototype/student-parent-app/index.html"),
    (2, "Academy Staff App", "prototype/academy-app/index.html"),
    (3, "Academy Dashboard (Web)", "prototype/academy-dashboard/index.html"),
    (4, "Super Admin Panel (Web)", "prototype/super-admin/index.html"),
]
for a in apps:
    row += 1
    for i, v in enumerate(a, 1):
        ws.cell(row=row, column=i, value=v)
    style_data_row(ws, row, 3)

row += 2
ws.cell(row=row, column=1, value="Tester Name:").font = BOLD_FONT
row += 1
ws.cell(row=row, column=1, value="Test Date:").font = BOLD_FONT
row += 1
ws.cell(row=row, column=1, value="Overall Result:").font = BOLD_FONT
ws.cell(row=row, column=2, value="").font = NORMAL_FONT
row += 1
ws.cell(row=row, column=1, value="Total Issues Found:").font = BOLD_FONT


# ══════════════════════════════════════════════
# SHEET 2: SCREEN INVENTORY
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("A - Screen Inventory")
ws2.sheet_properties.tabColor = "2E86C1"
headers = ["#", "App", "Screen/Page", "Navigation Path", "Status", "Tester", "Notes / Issues", "Severity", "Screenshot"]
set_col_widths(ws2, [5, 22, 22, 35, 10, 12, 40, 12, 15])

row = 1
for i, h in enumerate(headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, len(headers))

ws2.add_data_validation(STATUS_VALIDATION)
ws2.add_data_validation(SEVERITY_VALIDATION)

# Student & Parent App screens
screens = {
    "A1. Student & Parent App (21 screens)": [
        (1, "Splash", "App launch"),
        (2, "Login", "Tap splash screen"),
        (3, "OTP", "Login -> Get OTP"),
        (4, "Student Home", "OTP verify (as Student)"),
        (5, "Parent Home", "OTP verify (as Parent)"),
        (6, "Learn (Curriculum)", "Bottom nav -> Learn"),
        (7, "Skill Detail", "Learn -> tap any skill"),
        (8, "Feed", "Bottom nav -> Feed"),
        (9, "Play (Assessments)", "Bottom nav -> Play"),
        (10, "Parent Cricket", "Play -> Parent Cricket (parent mode)"),
        (11, "More Menu", "Bottom nav -> More"),
        (12, "Profile", "More -> Profile"),
        (13, "Notifications", "More -> Notifications"),
        (14, "Attendance History", "More -> Attendance History"),
        (15, "Settings", "More -> Settings"),
        (16, "Academy Search", "Home -> Find Academy"),
        (17, "Academy Profile", "Academy Search -> tap academy"),
        (18, "Apply (Admission)", "Academy Profile -> Apply"),
        (19, "Apply Status", "Submit application"),
        (20, "Matches", "More -> Matches"),
        (21, "Match Detail", "Matches -> tap any match"),
    ],
    "A2. Academy Staff App (29 screens)": [
        (1, "Splash", "App launch"),
        (2, "Login", "Tap splash screen"),
        (3, "OTP", "Login -> Get OTP"),
        (4, "Setup Wizard 1", "First login"),
        (5, "Setup Wizard 2", "Wizard -> Next"),
        (6, "Setup Wizard 3", "Wizard -> Next"),
        (7, "Setup Wizard 4", "Wizard -> Next"),
        (8, "Home", "OTP verify / Setup complete"),
        (9, "Students List", "Bottom nav -> Students"),
        (10, "Student Detail", "Students -> tap student"),
        (11, "Score Skill", "Student Detail -> Score Skill"),
        (12, "Add Student", "Students -> FAB (+)"),
        (13, "Promote Student", "Student Detail -> Promote"),
        (14, "Attendance", "Bottom nav -> Attend"),
        (15, "Attendance Report", "Attendance -> View Monthly Report"),
        (16, "Feed", "Bottom nav -> Feed"),
        (17, "Create Post", "Feed -> Create Post area"),
        (18, "More Menu", "Bottom nav -> More"),
        (19, "Curriculum Builder", "More -> Curriculum Builder"),
        (20, "Fee Management", "More -> Fee Management"),
        (21, "Staff Chat", "More -> Staff Chat"),
        (22, "Chat Detail", "Staff Chat -> tap channel"),
        (23, "Admissions", "More -> Admissions"),
        (24, "Staff Management", "More -> Staff Management"),
        (25, "Settings", "More -> Academy Settings"),
        (26, "Notifications", "Home -> bell icon"),
        (27, "Matches", "More -> Matches"),
        (28, "Match Detail", "Matches -> tap match"),
        (29, "Create Match", "Matches -> New Match"),
    ],
    "A3. Academy Dashboard - Web (12 pages)": [
        (1, "Dashboard", "Sidebar -> Dashboard"),
        (2, "Students", "Sidebar -> Students"),
        (3, "Curriculum", "Sidebar -> Curriculum"),
        (4, "Assessments", "Sidebar -> Assessments"),
        (5, "Attendance", "Sidebar -> Attendance"),
        (6, "Fee Management", "Sidebar -> Fee Management"),
        (7, "Staff", "Sidebar -> Staff"),
        (8, "Matches", "Sidebar -> Matches"),
        (9, "Feed & Posts", "Sidebar -> Feed & Posts"),
        (10, "Parent Community", "Sidebar -> Parent Community"),
        (11, "Reports", "Sidebar -> Reports"),
        (12, "Settings", "Sidebar -> Settings"),
    ],
    "A4. Super Admin Panel - Web (9 pages)": [
        (1, "Dashboard", "Sidebar -> Dashboard"),
        (2, "Academies", "Sidebar -> Academies"),
        (3, "Content Library", "Sidebar -> Content Library"),
        (4, "Users", "Sidebar -> Users"),
        (5, "Subscriptions & Billing", "Sidebar -> Subscriptions"),
        (6, "Tournaments", "Sidebar -> Tournaments"),
        (7, "Reports & Analytics", "Sidebar -> Reports"),
        (8, "Support Tickets", "Sidebar -> Support Tickets"),
        (9, "System Config", "Sidebar -> System Config"),
    ],
}

global_num = 0
for section_name, screen_list in screens.items():
    row += 1
    add_section_row(ws2, row, section_name, len(headers))
    app_name = section_name.split(".")[1].strip().split("(")[0].strip()
    for num, screen, nav in screen_list:
        row += 1
        global_num += 1
        ws2.cell(row=row, column=1, value=global_num)
        ws2.cell(row=row, column=2, value=app_name)
        ws2.cell(row=row, column=3, value=screen)
        ws2.cell(row=row, column=4, value=nav)
        # Status column - add validation
        STATUS_VALIDATION.add(ws2.cell(row=row, column=5))
        ws2.cell(row=row, column=5, value="")
        ws2.cell(row=row, column=6, value="")  # Tester
        ws2.cell(row=row, column=7, value="")  # Notes
        SEVERITY_VALIDATION.add(ws2.cell(row=row, column=8))
        ws2.cell(row=row, column=8, value="")  # Severity
        ws2.cell(row=row, column=9, value="")  # Screenshot
        style_data_row(ws2, row, len(headers))

ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row}"
ws2.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SHEET 3: WORKFLOW TESTS
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("B - Workflow Tests")
ws3.sheet_properties.tabColor = "27AE60"
wf_headers = ["WF#", "Step", "App", "Screen", "Action", "Expected Result", "Status", "Actual Result", "Tester", "Severity", "Priority", "Notes / Bug Description"]
set_col_widths(ws3, [6, 6, 18, 22, 40, 40, 10, 35, 12, 12, 16, 40])

row = 1
for i, h in enumerate(wf_headers, 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, len(wf_headers))

ws3.add_data_validation(STATUS_VALIDATION)
ws3.add_data_validation(SEVERITY_VALIDATION)
ws3.add_data_validation(PRIORITY_VALIDATION)

workflows = [
    {
        "id": "WF-01",
        "name": "WORKFLOW 1: Match Creation -> Availability -> Fee Collection -> Finance",
        "scenario": "Coach creates a weekend match. Parents/students confirm availability. Coach collects fees. Collection reflects in finance.",
        "steps": [
            ("Academy Staff", "More -> Matches", 'Tap "New Match"', "Create Match screen opens"),
            ("Academy Staff", "Create Match", 'Fill: Title="Weekend Practice Match", Type=Internal, Format=T10, Date, Time, Venue, Fee=Rs.200, Who=All Batches', "All fields accept input"),
            ("Academy Staff", "Create Match", 'Tap "Create Match & Notify Players"', 'Toast: "Match created! Notifications sent" -> Navigate to Matches list'),
            ("Academy Staff", "Matches", "Verify new match appears", "Match card shows: status=Collecting Availability, 0/X available, fee info"),
            ("Student App", "Home", "Check for Match Invite card", 'Match Invite section appears with match title, date, fee, Available/Can\'t/Maybe buttons'),
            ("Student App", "Home", 'Tap "Available" button', 'Toast: "Marked as Available!"'),
            ("Student App", "More -> Matches", "Verify match status", 'My Status = "You\'re In!", fee shows "Rs.200 Due"'),
            ("Student App", "Match Detail", "Tap the match card", 'Match detail opens with: info card, "You\'re confirmed" banner, fee section with payment options'),
            ("Student App", "Match Detail", "Verify fee payment options", "GPay, PhonePe, Cash at Academy buttons visible"),
            ("Academy Staff", "Match Detail", "Open same match", "Available count increased, student shows in Available list"),
            ("Academy Staff", "Match Detail", "Check Fee Collection section", "Progress bar shows X/Y paid, unpaid players listed"),
            ("Academy Staff", "Match Detail", 'Tap "Record" on unpaid player', "Record Payment modal opens with: student name, amount=Rs.200, match title, payment methods"),
            ("Academy Staff", "Payment Modal", 'Select "GPay - Mobile 1", enter reference, tap "Record Payment"', 'Toast: "Payment of Rs.200 recorded for [student]"'),
            ("Academy Staff", "Match Detail", "Verify fee progress updated", "Paid count increases, progress bar advances"),
            ("Academy Staff", "Match Detail", 'Tap "Announce Teams"', 'Toast: "Teams finalized & announced!"'),
            ("Student App", "Match Detail", "Refresh/recheck", 'Status changes to "Teams Announced", team list visible with "You" highlighted'),
            ("Academy Staff", "Match Detail", 'Tap "Start Match"', 'Toast: "Match started!"'),
            ("Academy Staff", "Match Detail", 'Tap "Share to Feed"', 'Toast: "Match posted to academy feed!"'),
            ("Academy Staff", "Fee Management", "Navigate to fee management", "Match fee collection reflects in overall stats (Collected amount)"),
            ("Academy Dashboard", "Matches", "Open Matches page", "Match card visible with availability bar, fee collection progress"),
            ("Academy Dashboard", "Matches", "Verify fee amounts", "Collected/Pending amounts match the payments recorded"),
            ("Academy Dashboard", "Fee Management", "Open Fee Management", "Revenue stats should include match fee collections"),
            ("Academy Dashboard", "Dashboard", "Check Dashboard", "Revenue chart and stats reflect updated collections"),
        ],
    },
    {
        "id": "WF-02",
        "name": "WORKFLOW 2: Student Admission -> Enrollment -> Stage Assignment",
        "scenario": "New student finds academy, applies for admission. Academy reviews and accepts. Student gets enrolled with stage assignment.",
        "steps": [
            ("Student App", "Home", 'Tap "Find Academy" quick action', "Academy Search screen opens"),
            ("Student App", "Academy Search", "View nearby academies", "3 academy cards shown with name, rating, distance, student count"),
            ("Student App", "Academy Search", 'Tap "SAM Cricket Academy"', "Academy Profile opens with details, batches, fee info"),
            ("Student App", "Academy Profile", 'Tap "Apply for Admission"', "Admission form opens with pre-filled fields"),
            ("Student App", "Apply", "Fill form: Name, Age, Experience, Preferred Batch, Parent details", "All fields accept input"),
            ("Student App", "Apply", 'Tap "Submit Application"', 'Apply Status screen: "Application Submitted!", "Under Review" status'),
            ("Academy Staff", "Home", "Check alerts section", '"Pending Admissions" alert card shows count increased'),
            ("Academy Staff", "More -> Admissions", "Open Admissions", "New applicant card visible with details"),
            ("Academy Staff", "Admissions", 'Tap "Accept" on applicant', "Modal opens with Role, Stage, Batch selection dropdowns"),
            ("Academy Staff", "Admissions Modal", 'Select Role=Batsman, Stage=1, Batch=Morning U-12, tap "Confirm"', 'Toast: "Student accepted"'),
            ("Academy Staff", "Students", "Navigate to Students list", "New student appears in the list with assigned Stage and Batch"),
            ("Academy Dashboard", "Dashboard", "Check dashboard", "Total Students count should reflect new addition"),
            ("Academy Dashboard", "Students", "Open Students page", "New student visible in table with assigned details"),
            ("Super Admin", "Dashboard", "Check platform stats", "Total Students count increases across platform"),
        ],
    },
    {
        "id": "WF-03",
        "name": "WORKFLOW 3: Student Progress -> Assessment -> Skill Scoring -> Promotion",
        "scenario": "Student learns skills, coach scores assessment, student passes and gets promoted.",
        "steps": [
            ("Student App", "Home", 'View "Next Up" skill card', 'Shows current in-progress skill (e.g. "Cut Shot - 60%")'),
            ("Student App", "Learn", "Tap bottom nav -> Learn", "Curriculum shows 4 categories with skills, progress %"),
            ("Student App", "Learn", 'Tap "Cut Shot" (in-progress)', "Skill Detail opens with: technique video, drills, assessment result"),
            ("Student App", "Skill Detail", "Verify content sections", "Technique Video, Coach Demo, 4 Drill Instructions, Latest Assessment"),
            ("Student App", "Play", "Tap bottom nav -> Play", "Upcoming Assessments tab shows scheduled assessment"),
            ("Student App", "Play", 'Tap "Results" tab', "Assessment Results shown with scores, grades, coach notes"),
            ("Academy Staff", "Students", "Tap on the student", "Student Detail opens with Progress/Attendance/Notes tabs"),
            ("Academy Staff", "Student Detail", 'Tap "Score Skill"', "Score Skill screen opens with drill slider and coach assessment slider"),
            ("Academy Staff", "Score Skill", "Set Drill Score=13/15, Coach Assessment=8/10", "Total score auto-calculates (weighted 60%+40%) and displays"),
            ("Academy Staff", "Score Skill", 'Add Coach Notes, tap "Save Score"', "Toast: score saved, navigates back"),
            ("Student App", "Play -> Results", "Check results", "New assessment result appears with score and coach note"),
            ("Academy Staff", "Student Detail", "Check if promotion-ready", 'If all skills passed, "Promote" button appears'),
            ("Academy Staff", "Promote Student", 'Tap "Promote"', "Promotion screen with readiness checklist, stage progression"),
            ("Academy Staff", "Promote Student", 'Add coach remarks, tap "Approve Promotion"', 'Toast: "Promoted to Stage X+1!"'),
            ("Student App", "Home", "Check stage display", "Student stage should update (e.g. Stage 3 -> Stage 4)"),
            ("Student App", "Feed", "Check feed", 'System achievement post: "Achievement Unlocked!" with promotion details'),
            ("Academy Dashboard", "Assessments", "Check Promotion Queue tab", "Student should move out of promotion queue"),
            ("Academy Dashboard", "Dashboard", "Check Stage Distribution chart", "Doughnut chart reflects updated distribution"),
        ],
    },
    {
        "id": "WF-04",
        "name": "WORKFLOW 4: Attendance Marking -> Low Attendance Alert -> Parent Notification",
        "scenario": "Staff marks daily attendance. A student has low attendance. System flags it. Parent gets notified.",
        "steps": [
            ("Academy Staff", "Home", 'Tap "Mark Attendance" quick action', "Attendance screen opens with today's date"),
            ("Academy Staff", "Attendance", "Verify batch filter", "Batch chips: All, Morning U-12, Evening U-14, Weekend All Ages"),
            ("Academy Staff", "Attendance", "Tap check for 8 students", "Buttons turn solid green, Present count = 8"),
            ("Academy Staff", "Attendance", "Tap X for 1 student", "Button turns solid red, Absent count = 1"),
            ("Academy Staff", "Attendance", "Tap clock for 1 student", "Button turns solid amber, Late count = 1"),
            ("Academy Staff", "Attendance", 'Tap "Mark All Present"', 'All buttons turn green, toast: "All students marked present!"'),
            ("Academy Staff", "Attendance", "Toggle - tap green button again", "Button deselects, Present count decreases"),
            ("Academy Staff", "Attendance", 'Tap "Submit Attendance"', 'Toast: "Attendance submitted successfully!"'),
            ("Academy Staff", "Attendance", 'Tap "View Monthly Report"', "Attendance Report: avg%, batch breakdown, low attendance alerts"),
            ("Academy Staff", "Attendance Report", 'Check "Low Attendance" section', "Students below 65% flagged with name, batch, percentage"),
            ("Student App", "Home", "Check attendance section", "Monthly attendance % and calendar grid visible"),
            ("Student App", "More -> Attendance", "Open Attendance History", "Calendar view: green (present), red (absent), empty (holiday)"),
            ("Academy Dashboard", "Attendance", "Open Attendance page", "Present/Absent/Late stat cards, weekly grid, monthly heatmap"),
            ("Academy Dashboard", "Attendance", 'Click "Mark Today"', "Modal opens with student list and status dropdowns"),
            ("Academy Dashboard", "Attendance", 'Check "Low Attendance Alerts"', 'Students below threshold listed with "Notify Parent" button'),
            ("Academy Dashboard", "Attendance", 'Click "Notify Parent"', 'Toast: "Reminder sent to [parent name]"'),
        ],
    },
    {
        "id": "WF-05",
        "name": "WORKFLOW 5: Fee Management -> Monthly Due -> Payment Recording -> Collection Tracking",
        "scenario": "Monthly fees are due. Staff records payments via different methods. Finance reflects collections.",
        "steps": [
            ("Academy Staff", "More -> Fee Mgmt", "Open Fee Management", "Stats: Collected, Pending, Total amounts shown"),
            ("Academy Staff", "Fee Management", "Check month filter", "Chips: All Months, Jan 2026, Feb 2026, Mar 2026"),
            ("Academy Staff", "Fee Management", "Verify pending students", "Each student shows: name, batch, month-wise breakdown"),
            ("Academy Staff", "Fee Management", "Check Vikram T (3 months due)", "Shows Jan, Feb, Mar 2026 rows with Rs.3,500 each, total Rs.10,500"),
            ("Academy Staff", "Fee Management", 'Verify "Oldest" tag', 'Jan 2026 row tagged as "Oldest" in red'),
            ("Academy Staff", "Fee Management", 'Tap "Record" on Jan 2026 row', "Payment modal: student name, month=Jan 2026, amount=Rs.3,500"),
            ("Academy Staff", "Payment Modal", "Verify payment methods", "6 options: GPay Mobile 1, GPay Mobile 2, Cash, Bank Transfer, PhonePe, Card/UPI"),
            ("Academy Staff", "Payment Modal", 'Select "Cash", leave reference empty', "Cash button highlights with accent border"),
            ("Academy Staff", "Payment Modal", 'Tap "Record Payment"', 'Toast: "Payment of Rs.3,500 recorded for Vikram T"'),
            ("Academy Staff", "Fee Management", 'Tap "Record All" on Deepak R', 'Payment modal: month="All (2 months)", amount=Rs.8,000'),
            ("Academy Staff", "Fee Management", 'Tap "Send Reminder" on a student', 'Toast: "Reminder sent to [student]"'),
            ("Academy Staff", "Fee Management", 'Tap "Send Reminder to All"', 'Toast: "Reminders sent to X students!" with correct count'),
            ("Academy Dashboard", "Fee Management", "Open Fee Management", "Summary boxes: Collected, Pending, Total match expected values"),
            ("Academy Dashboard", "Fee Management", "Check pending table", "Students with Due/Partial status listed"),
            ("Academy Dashboard", "Fee Management", 'Click "Record" on a student', 'Toast: "Payment recorded"'),
            ("Academy Dashboard", "Fee Management", 'Click "Send Bulk Reminder"', 'Toast: "Bulk reminders sent to X parents!"'),
            ("Academy Dashboard", "Dashboard", "Check revenue stats", "Revenue This Month and Pending Fees reflect correct values"),
            ("Student App (Parent)", "Parent Home", "Check fee section", 'Shows fee amount, due date, "Pay Now" button'),
            ("Super Admin", "Billing", "Check MRR", "Monthly Recurring Revenue reflects academy collections"),
        ],
    },
    {
        "id": "WF-06",
        "name": "WORKFLOW 6: Social Feed -> Post Creation -> Visibility Control",
        "scenario": "Coach creates posts with different visibility. Students/parents see only relevant posts.",
        "steps": [
            ("Academy Staff", "Feed", "Check existing posts", "4 posts visible with author, time, visibility badges"),
            ("Academy Staff", "Feed", "Tap create post area", "Create Post screen opens"),
            ("Academy Staff", "Create Post", 'Select "Public" visibility', "Public chip highlighted"),
            ("Academy Staff", "Create Post", 'Type message, tap "Post"', 'Toast: "Post published!", navigates to Feed'),
            ("Academy Staff", "Create Post", 'Select "Staff Only" visibility', "Staff Only chip highlighted"),
            ("Academy Staff", "Create Post", 'Tap "Photo" button', 'Toast: "Camera opened"'),
            ("Academy Staff", "Create Post", 'Tap "Video" button', 'Toast: "Video feature coming soon"'),
            ("Academy Staff", "Create Post", 'Tap "Match" button', "Navigates to Create Match screen"),
            ("Student App", "Feed", "Check feed", "Posts with Public and Academy visibility visible"),
            ("Student App", "Feed", 'Tap "Academy" tab', "Academy feed shows"),
            ("Student App", "Feed", "Tap heart on a post", 'Toast: "Liked!"'),
            ("Student App", "Feed", "Tap share on a post", 'Toast: "Shared!"'),
            ("Academy Dashboard", "Feed & Posts", "Open Feed page", "All posts visible with visibility badges, engagement stats"),
            ("Academy Dashboard", "Feed & Posts", 'Click "Create Post"', "Modal with visibility selector and content textarea"),
            ("Academy Dashboard", "Feed & Posts", "Check content moderation card", '"No flagged posts" with green shield'),
        ],
    },
    {
        "id": "WF-07",
        "name": "WORKFLOW 7: Curriculum Builder -> Skill Management -> Student Learning",
        "scenario": "Academy customizes curriculum by toggling skills. Students see updated curriculum.",
        "steps": [
            ("Academy Staff", "More -> Curriculum", "Open Curriculum Builder", "4 stages shown, each with skills list"),
            ("Academy Staff", "Curriculum", "Tap Stage 1 header", "Stage expands/collapses showing skills"),
            ("Academy Staff", "Curriculum", "Toggle a skill checkbox off", "Skill unchecked (visual feedback)"),
            ("Academy Staff", "Curriculum", 'Tap "Add Custom Skill"', 'Toast: "Custom skill editor in full app"'),
            ("Academy Staff", "Curriculum", 'Tap "Save Curriculum"', 'Toast: "Curriculum saved!"'),
            ("Student App", "Learn", "Open Learn tab", "4 categories shown with skill counts and progress"),
            ("Student App", "Learn", "Tap a category to expand", "Skills listed with status icons"),
            ("Student App", "Learn", "Tap a passed skill", "Skill Detail shows technique video, drills, assessment"),
            ("Student App", "Learn", "Tap a locked skill", "Should not navigate (opacity 0.5, unclickable)"),
            ("Academy Dashboard", "Curriculum", "Open Curriculum page", "Two-column: stage/skill tree + skill detail panel"),
            ("Academy Dashboard", "Curriculum", "Click a skill", "Right panel shows: video status, drill description, passing criteria"),
            ("Academy Dashboard", "Curriculum", "Toggle stage expand", "Stage group expands/collapses with chevron animation"),
        ],
    },
    {
        "id": "WF-08",
        "name": "WORKFLOW 8: Parent Match -> Availability -> Match Execution",
        "scenario": "Academy organizes Parents vs Students fun match. Parents respond. Match happens.",
        "steps": [
            ("Academy Staff", "Create Match", "Type=Parent vs Student, Format=T8, Fee=Rs.0", "All fields populated correctly"),
            ("Academy Staff", "Create Match", 'Tap "Create Match & Notify"', "Toast confirms, navigates to Matches"),
            ("Student App", "Home", "Check match invite", '"Parents vs Students Fun Match" card with "Free" badge'),
            ("Student App", "Match Detail", "Open match", 'Info card shows: Free event, eligibility=Students + Parents'),
            ("Student App", "Match Detail", 'Tap "Available"', "Button highlights green, toast confirms"),
            ("Student App", "Match Detail", 'Tap "Available" again', "Deselects (toggles back to none)"),
            ("Student App", "Match Detail", 'Tap "Maybe"', "Button highlights amber"),
            ("Student App (Parent)", "Parent Cricket", "Check parent cricket section", "Upcoming match visible with availability buttons"),
            ("Academy Staff", "Match Detail", "Check responses", "Available/Not Available/Maybe/Pending counts shown"),
            ("Academy Staff", "Match Detail", "Verify free match", 'No fee collection section (shows "Free Event" instead)'),
            ("Academy Dashboard", "Matches", "Check match card", "Free event badge, no fee progress bar"),
            ("Academy Dashboard", "Parent Community", "Check upcoming match", "Match details visible in Parent Community page"),
        ],
    },
    {
        "id": "WF-09",
        "name": "WORKFLOW 9: Staff Management -> Communication -> Chat",
        "scenario": "Multiple coaches communicate via staff chat. Head coach assigns tasks.",
        "steps": [
            ("Academy Staff", "More -> Staff Mgmt", "Open Staff Management", "3 staff cards: Venkat (Owner), Coach Ravi, Kumar"),
            ("Academy Staff", "Staff Management", 'Tap "Invite Staff"', 'Toast: "Invite link copied!"'),
            ("Academy Staff", "More -> Staff Chat", "Open Staff Chat", "5 channels listed with unread badges"),
            ("Academy Staff", "Staff Chat", 'Tap "All Staff" (3 unread)', "Chat Detail opens with message thread"),
            ("Academy Staff", "Chat Detail", "View messages", "5 messages with sender names, timestamps, sent/received styling"),
            ("Academy Staff", "Chat Detail", "Tap send button", 'Toast: "Message sent!"'),
            ("Academy Staff", "Chat Detail", "Tap attachment (+)", 'Toast: "Attachment options"'),
            ("Academy Dashboard", "Staff", "Open Staff page", "Staff cards with role, batches, phone, performance table"),
            ("Academy Dashboard", "Staff", 'Click "Invite Staff"', "Modal with name, role, phone, batch assignment"),
        ],
    },
    {
        "id": "WF-10",
        "name": "WORKFLOW 10: Platform Onboarding -> Academy Setup -> Trial",
        "scenario": "New academy signs up, goes through setup wizard, starts 14-day trial.",
        "steps": [
            ("Academy Staff", "Splash", "Tap screen", "Login screen opens"),
            ("Academy Staff", "Login", 'Enter phone, tap "Get OTP"', "OTP screen opens"),
            ("Academy Staff", "OTP", "Enter 4 digits", "Auto-progresses, verify button available"),
            ("Academy Staff", "OTP", 'Tap "Verify & Login"', 'Toast: "OTP verified!", navigates to Home'),
            ("Academy Staff", "Setup Wizard 1", "Fill: Academy Name, Description, Facilities", "All checkboxes and inputs functional"),
            ("Academy Staff", "Setup Wizard 1", 'Tap "Next"', "Step 2 opens (Batches)"),
            ("Academy Staff", "Setup Wizard 2", "Fill: Batch name, times, days, max students", "Day selector and inputs work"),
            ("Academy Staff", "Setup Wizard 2", 'Tap "Next"', "Step 3 opens (Attendance method)"),
            ("Academy Staff", "Setup Wizard 3", "Select attendance method", "Radio buttons functional (Staff/Self/Both)"),
            ("Academy Staff", "Setup Wizard 3", 'Tap "Next"', "Step 4 opens (Fee config)"),
            ("Academy Staff", "Setup Wizard 4", "Set monthly fee, due date, payment methods", "All inputs accept values"),
            ("Academy Staff", "Setup Wizard 4", 'Tap "Finish Setup"', 'Toast: "Academy setup complete!", navigates to Home'),
            ("Academy Staff", "Home", "Check trial badge", '"Trial: 12 days left" badge visible'),
            ("Academy Staff", "More", "Check subscription section", "Trial status shown with days remaining"),
            ("Academy Staff", "Settings", "Check subscription", 'Trial progress bar and "Upgrade Now" button'),
            ("Super Admin", "Dashboard", 'Check "Recent Signups"', "New academy should appear in list"),
            ("Super Admin", "Academies", "Check academy list", 'New academy with "Trial" status badge'),
        ],
    },
    {
        "id": "WF-11",
        "name": "WORKFLOW 11: Assessment Scheduling -> Exam -> Results",
        "scenario": "Coach creates assessment. Students take it. Results are recorded and visible.",
        "steps": [
            ("Academy Dashboard", "Assessments", 'Click "Create Assessment"', "Modal with title, date, stage, batch, coach, skills fields"),
            ("Academy Dashboard", "Assessments", "Fill and submit", 'Toast: "Assessment created!"'),
            ("Academy Dashboard", "Assessments", 'Check "Upcoming" tab', "New assessment visible with date, batch, student count"),
            ("Student App", "Play", 'Check "Upcoming" tab', "Assessment shown with skill, date, time, coach"),
            ("Academy Staff", "Student Detail", 'Tap "Score Skill"', "Score screen with drill slider (0-15) and assessment slider (1-10)"),
            ("Academy Staff", "Score Skill", "Move drill slider to 12", 'Display updates to "12/15"'),
            ("Academy Staff", "Score Skill", "Move assessment slider to 8", 'Display updates to "8/10"'),
            ("Academy Staff", "Score Skill", "Verify total", "Weighted calculation shown (60% drill + 40% assessment)"),
            ("Academy Staff", "Score Skill", 'Tap "Save Score"', "Toast confirms, navigates back"),
            ("Student App", "Play -> Results", "Check results", "New result with score, grade, coach note"),
            ("Academy Dashboard", "Assessments", 'Check "Recent Results" tab', "Assessment result in table with scores"),
            ("Academy Dashboard", "Assessments", 'Check "Promotion Queue" tab', "If student qualifies (score>=85, attendance>=90), appears here"),
        ],
    },
    {
        "id": "WF-12",
        "name": "WORKFLOW 12: Super Admin Content Management",
        "scenario": "Platform team uploads coaching content to shared library for academies.",
        "steps": [
            ("Super Admin", "Content Library", "Open page", "Stats: Created (72), In Review (12), Published (0), Remaining (446)"),
            ("Super Admin", "Content Library", 'Click "Batting" tab', "Grid filters to batting content only"),
            ("Super Admin", "Content Library", 'Click "Bowling" tab', "Shows filtered content (may be empty)"),
            ("Super Admin", "Content Library", "Use Stage dropdown", "Filters by stage (1-4)"),
            ("Super Admin", "Content Library", "Use Status dropdown", "Filters by Draft/Review/Published"),
            ("Super Admin", "Content Library", 'Click "Upload New Content"', "Modal: title, skill, category, stage, file upload zone, description"),
            ("Super Admin", "Content Library", "Click drag-drop zone", 'Toast: "File picker would open here"'),
            ("Super Admin", "Content Library", 'Fill form and click "Upload"', 'Toast: "Content uploaded!", modal closes'),
            ("Super Admin", "Content Library", "Click a content card", 'Toast: "Opening content: [title]"'),
        ],
    },
    {
        "id": "WF-13",
        "name": "WORKFLOW 13: Support Ticket Lifecycle",
        "scenario": "Academy reports issue. Support team tracks and resolves it.",
        "steps": [
            ("Super Admin", "Support Tickets", "Check stats", "Open, In Progress, Resolved counts shown"),
            ("Super Admin", "Support Tickets", 'Click "Open" tab', "Filters to open tickets only"),
            ("Super Admin", "Support Tickets", 'Click "Create Ticket"', "Modal: academy dropdown, subject, priority, assign to, description"),
            ("Super Admin", "Support Tickets", "Fill and submit", 'Toast: "Ticket created!", modal closes'),
            ("Super Admin", "Support Tickets", "Click view (eye icon) on ticket", "Ticket detail modal: subject, priority, status, academy, description"),
            ("Super Admin", "Ticket Detail", 'Click "Reply"', 'Toast: "Reply sent"'),
            ("Super Admin", "Ticket Detail", 'Click "Mark Resolved"', "Toast confirms, modal closes"),
            ("Super Admin", "Ticket Detail", 'Click "Reassign"', 'Toast: "Ticket reassigned"'),
        ],
    },
]

for wf in workflows:
    row += 1
    add_section_row(ws3, row, f'{wf["id"]}: {wf["name"]}', len(wf_headers))
    row += 1
    add_sub_header_row(ws3, row, f'Scenario: {wf["scenario"]}', len(wf_headers))
    for step_num, (app, screen, action, expected) in enumerate(wf["steps"], 1):
        row += 1
        ws3.cell(row=row, column=1, value=wf["id"])
        ws3.cell(row=row, column=2, value=step_num)
        ws3.cell(row=row, column=3, value=app)
        ws3.cell(row=row, column=4, value=screen)
        ws3.cell(row=row, column=5, value=action)
        ws3.cell(row=row, column=6, value=expected)
        STATUS_VALIDATION.add(ws3.cell(row=row, column=7))
        ws3.cell(row=row, column=7, value="")
        ws3.cell(row=row, column=8, value="")  # Actual Result
        ws3.cell(row=row, column=9, value="")  # Tester
        SEVERITY_VALIDATION.add(ws3.cell(row=row, column=10))
        ws3.cell(row=row, column=10, value="")
        PRIORITY_VALIDATION.add(ws3.cell(row=row, column=11))
        ws3.cell(row=row, column=11, value="")
        ws3.cell(row=row, column=12, value="")  # Notes
        style_data_row(ws3, row, len(wf_headers))

ws3.auto_filter.ref = f"A1:{get_column_letter(len(wf_headers))}{row}"
ws3.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SHEET 4: UI/UX CHECKS
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("C - UI UX Checks")
ws4.sheet_properties.tabColor = "E74C3C"
ui_headers = ["#", "Category", "Test", "App", "Status", "Tester", "Notes", "Severity"]
set_col_widths(ws4, [5, 20, 45, 25, 10, 12, 40, 12])

row = 1
for i, h in enumerate(ui_headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, len(ui_headers))

ws4.add_data_validation(STATUS_VALIDATION)
ws4.add_data_validation(SEVERITY_VALIDATION)

ui_tests = {
    "C1. Theme Toggle": [
        (1, "Toggle to dark mode", "Student App - Settings -> Dark Mode"),
        (2, "Toggle to dark mode", "Academy Staff - (if available)"),
        (3, "Toggle to dark mode", "Academy Dashboard - top bar moon icon"),
        (4, "Toggle to dark mode", "Super Admin - top bar moon icon"),
        (5, "Charts reinitialize on theme change", "Dashboard (web)"),
        (6, "Theme persists on page reload", "All web apps (localStorage)"),
    ],
    "C2. Responsive/Mobile (Web Apps)": [
        (7, "Sidebar collapses on mobile width", "Academy Dashboard"),
        (8, "Hamburger menu works", "Academy Dashboard"),
        (9, "Tables are scrollable on small screens", "Academy Dashboard"),
        (10, "Charts resize properly", "Academy Dashboard"),
        (11, "Sidebar overlay works", "Super Admin"),
    ],
    "C3. Navigation": [
        (12, "Back button works on all sub-screens", "Academy Staff App"),
        (13, "Back button works on all sub-screens", "Student App"),
        (14, "Bottom nav highlights correct tab", "Both mobile apps"),
        (15, "Sidebar highlights active page", "Both web apps"),
        (16, "Tab switching works (Play, Student Detail, Feed)", "Student App"),
        (17, "Filter chips toggle correctly (single-select)", "Both mobile apps"),
    ],
}

for section_name, tests in ui_tests.items():
    row += 1
    add_section_row(ws4, row, section_name, len(ui_headers))
    for num, test, app in tests:
        row += 1
        ws4.cell(row=row, column=1, value=num)
        ws4.cell(row=row, column=2, value=section_name.split(".")[1].strip())
        ws4.cell(row=row, column=3, value=test)
        ws4.cell(row=row, column=4, value=app)
        STATUS_VALIDATION.add(ws4.cell(row=row, column=5))
        ws4.cell(row=row, column=6, value="")
        ws4.cell(row=row, column=7, value="")
        SEVERITY_VALIDATION.add(ws4.cell(row=row, column=8))
        style_data_row(ws4, row, len(ui_headers))

# Charts section
row += 1
add_section_row(ws4, row, "C4. Charts (Web Apps)", len(ui_headers))
charts = [
    (1, "Enrollment trend line chart", "Dashboard - Dashboard"),
    (2, "Revenue bar chart", "Dashboard - Dashboard"),
    (3, "Stage distribution doughnut", "Dashboard - Dashboard"),
    (4, "Role distribution doughnut", "Dashboard - Dashboard"),
    (5, "Fee collection trend", "Dashboard - Fee Management"),
    (6, "Progress report line chart", "Dashboard - Reports"),
    (7, "Attendance report bar chart", "Dashboard - Reports"),
    (8, "Financial report bar chart", "Dashboard - Reports"),
    (9, "Academy growth line chart", "Super Admin - Dashboard"),
    (10, "Revenue growth bar chart", "Super Admin - Dashboard"),
    (11, "Plan distribution pie chart", "Super Admin - Dashboard"),
    (12, "Billing revenue line chart", "Super Admin - Billing"),
    (13, "User growth line chart", "Super Admin - Reports"),
    (14, "Feature usage bar chart", "Super Admin - Reports"),
    (15, "Churn rate line chart", "Super Admin - Reports"),
    (16, "Charts do NOT infinitely expand height", "All web apps - All pages"),
]
for num, test, app in charts:
    row += 1
    ws4.cell(row=row, column=1, value=num)
    ws4.cell(row=row, column=2, value="Charts")
    ws4.cell(row=row, column=3, value=test)
    ws4.cell(row=row, column=4, value=app)
    STATUS_VALIDATION.add(ws4.cell(row=row, column=5))
    ws4.cell(row=row, column=6, value="")
    ws4.cell(row=row, column=7, value="")
    SEVERITY_VALIDATION.add(ws4.cell(row=row, column=8))
    style_data_row(ws4, row, len(ui_headers))

ws4.auto_filter.ref = f"A1:{get_column_letter(len(ui_headers))}{row}"
ws4.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SHEET 5: DATA CONSISTENCY
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("D - Data Consistency")
ws5.sheet_properties.tabColor = "8E44AD"
dc_headers = ["#", "Data Point", "Academy Staff App", "Student App", "Academy Dashboard", "Super Admin", "Status", "Tester", "Notes"]
set_col_widths(ws5, [5, 35, 22, 22, 22, 22, 10, 12, 35])

row = 1
for i, h in enumerate(dc_headers, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, len(dc_headers))

ws5.add_data_validation(STATUS_VALIDATION)

data_checks = [
    (1, 'Academy name = "SAM Cricket Academy"', "Home hero card", "Home header", "Top bar title", "Academies list"),
    (2, 'Head Coach = "Venkat"', "More menu owner", "Feed posts author", "Admin profile label", "-"),
    (3, "Total students = 85", "Home hero", "-", "Dashboard stat", "Platform stats"),
    (4, 'Student "Ravi Kumar" = Stage 3', "Students list", "Home stage display", "Students table", "-"),
    (5, "Monthly fee = Rs.3,500", "Fee Management", "Parent Home fee card", "Settings fee config", "-"),
    (6, "3 batches exist", "Attendance filters", "-", "Dashboard stat", "-"),
    (7, "Match fee = Rs.200 (Weekend match)", "Match Detail", "Match Detail", "Matches page", "-"),
    (8, "Vikram T = 3 months overdue", "Fee Management", "-", "Fee Management", "-"),
    (9, "Trial = 12 days left", "More menu badge", "-", "Settings subscription", "Academies status"),
]

for num, data_point, staff, student, dashboard, admin in data_checks:
    row += 1
    ws5.cell(row=row, column=1, value=num)
    ws5.cell(row=row, column=2, value=data_point)
    ws5.cell(row=row, column=3, value=staff)
    ws5.cell(row=row, column=4, value=student)
    ws5.cell(row=row, column=5, value=dashboard)
    ws5.cell(row=row, column=6, value=admin)
    STATUS_VALIDATION.add(ws5.cell(row=row, column=7))
    ws5.cell(row=row, column=8, value="")
    ws5.cell(row=row, column=9, value="")
    style_data_row(ws5, row, len(dc_headers))

ws5.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SHEET 6: KNOWN LIMITATIONS
# ══════════════════════════════════════════════
ws6 = wb.create_sheet("E - Known Limitations")
ws6.sheet_properties.tabColor = "F39C12"
kl_headers = ["#", "Limitation", "Affected Area", "Impact", "Notes"]
set_col_widths(ws6, [5, 55, 30, 15, 40])

row = 1
for i, h in enumerate(kl_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, len(kl_headers))

limitations = [
    (1, "Data is hardcoded - no real persistence between sessions", "All apps", "Expected"),
    (2, "State resets on page refresh", "All apps", "Expected"),
    (3, "Filter chips are visual-only in some places (don't actually filter)", "Fee month filter, some batch filters", "Expected"),
    (4, '"Coming soon" toasts on features not yet prototyped', "Various screens", "Expected"),
    (5, "No actual file upload", "Create Post photos/video, Super Admin content upload", "Expected"),
    (6, "No real-time sync between apps", "Cross-app workflows", "Expected"),
    (7, "OTP accepts any input", "Login flow", "Expected"),
    (8, "Charts use static data, not reactive to actions", "Web dashboards", "Expected"),
    (9, 'Parent Cricket team names still show "Chennai Dads XI"', "Student App parent cricket page", "Expected"),
    (10, "Academy search shows generic academies, not linked to enrolled", "Student App academy search", "Expected"),
]

for num, limitation, area, impact in limitations:
    row += 1
    ws6.cell(row=row, column=1, value=num)
    ws6.cell(row=row, column=2, value=limitation)
    ws6.cell(row=row, column=3, value=area)
    ws6.cell(row=row, column=4, value=impact)
    ws6.cell(row=row, column=5, value="")
    style_data_row(ws6, row, len(kl_headers))

ws6.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SHEET 7: BUG TRACKER
# ══════════════════════════════════════════════
ws7 = wb.create_sheet("Bug Tracker")
ws7.sheet_properties.tabColor = "C0392B"
bug_headers = ["Bug ID", "Date Found", "Tester", "App", "Screen", "Workflow#", "Step#",
               "Summary", "Steps to Reproduce", "Expected", "Actual", "Severity", "Priority",
               "Status", "Assigned To", "Fix Date", "Retest Status", "Screenshot/Video", "Comments"]
set_col_widths(ws7, [10, 12, 12, 18, 18, 10, 8, 35, 40, 30, 30, 12, 16, 12, 14, 12, 14, 18, 35])

row = 1
for i, h in enumerate(bug_headers, 1):
    ws7.cell(row=row, column=i, value=h)
style_header_row(ws7, row, len(bug_headers))

ws7.add_data_validation(SEVERITY_VALIDATION)
ws7.add_data_validation(PRIORITY_VALIDATION)

bug_status_val = DataValidation(type="list", formula1='"New,In Progress,Fixed,Verified,Reopened,Won\'t Fix,Deferred"', allow_blank=True)
ws7.add_data_validation(bug_status_val)
retest_val = DataValidation(type="list", formula1='"Not Tested,PASS,FAIL"', allow_blank=True)
ws7.add_data_validation(retest_val)

# Add 30 empty rows for bugs
for i in range(1, 31):
    row += 1
    ws7.cell(row=row, column=1, value=f"BUG-{i:03d}")
    for c in range(1, len(bug_headers) + 1):
        ws7.cell(row=row, column=c).border = THIN_BORDER
        ws7.cell(row=row, column=c).font = NORMAL_FONT
        ws7.cell(row=row, column=c).alignment = WRAP
    SEVERITY_VALIDATION.add(ws7.cell(row=row, column=12))
    PRIORITY_VALIDATION.add(ws7.cell(row=row, column=13))
    bug_status_val.add(ws7.cell(row=row, column=14))
    retest_val.add(ws7.cell(row=row, column=17))

ws7.auto_filter.ref = f"A1:{get_column_letter(len(bug_headers))}{row}"
ws7.freeze_panes = "A2"


# ══════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════
output_path = "d:/CricketApp/soruban-sports/docs/qa/soruban-sports-test-plan.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"Sheets: {len(wb.sheetnames)} — {', '.join(wb.sheetnames)}")
